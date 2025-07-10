import cv2
from pyzbar import pyzbar
import time
import socket
import csv
import threading
from tkinter import Tk, Canvas, Button, PhotoImage, Label
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Global variables
scanned_code = {}
output_path = 'output.csv'
Flag = True
product_count = 0
station_num = 0  # Initialize station_num

# Read station_num from config file
def read_config():
    global station_num
    file_path = 'config.txt'
    with open(file_path, 'r') as file:
        station_num = int(file.read().strip())

# Initialize Tkinter window and components
def initialize_gui():
    global window, canvas, Total_output, scanned_text, scan_frame

    window = Tk()
    window.geometry("1100x738")
    window.configure(bg="#FFFFFF")

    # Disable the close button
    window.protocol("WM_DELETE_WINDOW", disable_close_button)
    
    canvas = Canvas(
        window,
        bg="#FFFFFF",
        height=738,
        width=1100,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )
    canvas.place(x=0, y=0)

    # Create GUI elements (images, buttons, texts)
    # Replace 'image_1.png', 'image_2.png', 'image_3.png', 'button_1.png', 'button_2.png' with your actual image paths

    image_image_1 = PhotoImage(file="./image_1.png")
    canvas.create_image(550.0, 66.0, image=image_image_1)

    image_image_2 = PhotoImage(file="./image_2.png")
    canvas.create_image(805.0, 422.0, image=image_image_2)

    image_image_3 = PhotoImage(file="./image_3.png")
    canvas.create_image(301.0, 422.0, image=image_image_3)

    scan_frame = Canvas(master=window, height=70, width=200, background='#009999')
    scan_frame.place(x=850, y=30)
    scanned_text = Label(scan_frame, text='Scanned', font=("Inter", 30, "bold"), bg="#009999", fg='#009999')
    scanned_text.pack()

    canvas.create_text(
        135.0,
        284.0,
        anchor="nw",
        text="OUTPUT PRODUCED",
        fill="#000000",
        font=("Inter", 34 * -1)
    )

    canvas.create_text(
        629.0,
        275.0,
        anchor="nw",
        text="TARGETED OUTPUT",
        fill="#000000",
        font=("Inter", 34 * -1)
    )

    global Total_output
    Total_output = canvas.create_text(
        230.0,
        369.0,
        anchor="nw",
        text="0",
        fill="#000000",
        font=("Inter", 128 * -1)
    )

    canvas.create_text(
        718.0,
        369.0,
        anchor="nw",
        text="60",
        fill="#000000",
        font=("Inter", 128 * -1)
    )

    canvas.create_text(
        41.0,
        29.0,
        anchor="nw",
        text="SIEMENS",
        fill="#FFFFFF",
        font=("Inter", 64 * -1, "bold")
    )

    # Create other GUI elements as needed

    canvas.pack()

    button_image_1 = PhotoImage(file="./button_1.png")
    button_1 = Button(
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: print("button_1 clicked"),
        relief="flat"
    )
    button_1.place(x=225.0, y=637.0, width=142.0, height=59.0)

    button_image_2 = PhotoImage(file="./button_2.png")
    button_2 = Button(
        image=button_image_2,
        borderwidth=0,
        highlightthickness=0,
        command=stop_program,
        relief="flat"
    )
    button_2.place(x=728.0, y=637.0, width=142.0, height=59.0)

    # Start the GUI main loop
    update_gui(canvas, Total_output)
    window.mainloop()

def disable_close_button():
    pass  # Do nothing when the close button is pressed

def stop_program():
    global Flag, window

    Flag = False
    window.destroy()

    # Clear the contents of output.csv
    try:
        with open("output.csv", mode='w', newline='') as csv_file:
            csv_file.truncate(0)
        print("Contents of output.csv cleared")
    except Exception as e:
        print(f"Error clearing output.csv: {e}")

def update_gui(canvas, total_output):
    global product_count

    if not Flag:
        return

    canvas.itemconfig(total_output, text=str(product_count))
    canvas.after(1000, update_gui, canvas, total_output)  # Schedule the next update after 1 second

def change_scanned_text_color():
    scanned_text.config(fg="green")
    scanned_text.after(1000, lambda: scanned_text.config(fg="#009999"))  # Change color back after 1 second

# Function to save log entry to Excel
def save_to_excel(log_entry):
    today = datetime.now().strftime("%Y-%m-%d")
    output_xlsx_file = f"{today}_xl.xlsx"

    try:
        # Check if the Excel file exists
        if os.path.exists(output_xlsx_file):
            wb = load_workbook(output_xlsx_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        ws.append(log_entry)
        wb.save(output_xlsx_file)
        print(f"Log entry saved to {output_xlsx_file}")

    except Exception as e:
        print(f"Error saving log entry to Excel file: {e}")

# Function to scan barcode from webcam
def scan_barcode_from_webcam(cam_no):
    global scanned_code, product_count
    cap = cv2.VideoCapture(cam_no)
    try:
        while Flag:
            ret, frame = cap.read()
            if not ret:
                break
            barcodes = pyzbar.decode(frame)
            for barcode in barcodes:
                barcode_data = barcode.data.decode('utf-8')
                current_time = time.time()
                print("code scanned")

                change_scanned_text_color()  # Change the color when a code is scanned

                if barcode_data in scanned_code:
                    
                    if (scanned_code[barcode_data]['first_cam_scan']==cam_no) and (len(scanned_code[barcode_data]['cam_code'])==station_num+1):
                        del scanned_code[barcode_data]
                        print("dict reset")
                        continue
                    
                    if cam_no in scanned_code[barcode_data]['cam_code']:
                        continue
                    else:
                        scanned_code[barcode_data]['timestamps'].append(current_time)
                        scanned_code[barcode_data]['cam_code'].append(cam_no)
                        print(f"code scanned from cam: {cam_no}")

                        if len(scanned_code[barcode_data]['cam_code']) == station_num + 1:
                            product_count += 1
                            log_entry = [str(product_count)]
                            for i in range(station_num):
                                time_at_station = scanned_code[barcode_data]['timestamps'][i+1] - scanned_code[barcode_data]['timestamps'][i]
                                log_entry.append(str(time_at_station))

                            # Write log entry to file
                            with open(output_path, 'a', newline='') as outfile:
                                writer = csv.writer(outfile)
                                writer.writerow(log_entry)
                                print("output saved")
                            
                            # Save log entry to Excel
                            save_to_excel(log_entry)

                else:
                    scanned_code[barcode_data] = {'timestamps': [current_time], 'cam_code': [cam_no], 'first_cam_scan': cam_no}
                    print(f"code scanned from cam: {cam_no}")

    finally:
        cap.release()
        cv2.destroyAllWindows()

# Function to start server
def server_start():
    global station_num
    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    HOST = socket.gethostbyname(socket.gethostname())
    PORT = 9999
    server.bind((HOST, PORT))
    server.listen(5)
    print(f"Server started at {HOST}:{PORT}")

    clients = []  # List to store connected clients

    while Flag:
        try:
            server.settimeout(1)  # Set a timeout for server.accept() to periodically check the running flag
            communication_socket, address = server.accept()
            print(f"Connected to {address}")
            clients.append(communication_socket)  # Add the new client to the list

            # Send station_num to the new client
            communication_socket.send(str(station_num).encode('utf-8') + b'\n')

            # Send the Excel data to the new client
            today = datetime.now().strftime("%Y-%m-%d")
            output_xlsx_file = f"{today}_xl.xlsx"

            try:
                # Open the Excel file
                wb = load_workbook(output_xlsx_file)
                ws = wb.active

                # Iterate through the rows in the Excel file
                for row in ws.iter_rows(values_only=True):
                    response = ','.join(map(str, row)).encode('utf-8')
                    communication_socket.send(response + b'\n')  # Send newline to indicate end of row

                # Send end-of-transmission message
                communication_socket.send(b'ALL_ROWS_SENT\n')

            except FileNotFoundError:
                print(f"Excel file {output_xlsx_file} not found.")
                communication_socket.send(b'FILE_NOT_FOUND\n')

        except socket.timeout:
            pass  # Ignore timeout, continue loop
        except ConnectionResetError:
            print(f"Client {address} disconnected.")
            clients.remove(communication_socket)
        except Exception as e:
            print(f"Error: {e}")

    # Clean up remaining clients
    for client in clients:
        client.close()

    server.close()
    print("Server stopped.")

# Main function to start GUI and threads
if __name__ == "__main__":
    read_config()  # Read configuration

    # Start the GUI in the main thread
    initialize_gui_thread = threading.Thread(target=initialize_gui)
    initialize_gui_thread.start()

    # Start camera threads
    camera_threads = []
    for cam_no in range(station_num + 1):
        camera_thread = threading.Thread(target=scan_barcode_from_webcam, args=(cam_no,))
        camera_threads.append(camera_thread)
        camera_thread.start()

    # Start the server thread
    server_thread = threading.Thread(target=server_start)
    server_thread.start()

    # Wait for GUI initialization to complete
    initialize_gui_thread.join()

    # Start update GUI thread
    update_gui_thread = threading.Thread(target=update_gui, args=(canvas, Total_output))
    update_gui_thread.start()

    # Wait for all threads to finish
    for camera_thread in camera_threads:
        camera_thread.join()

    # Stop the server thread
    Flag = False
    server_thread.join()
    update_gui_thread.join()
